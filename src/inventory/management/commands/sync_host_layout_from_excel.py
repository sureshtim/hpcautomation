from datetime import datetime
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction

from inventory.models import Host


# ── Excel config ──────────────────────────────────────────────────────────────
EXCEL_FILENAME = "HPC_Not Retired_2-17-26.xlsx"
SHEET_NAME     = "Page 1"

# Lookup columns (tried in order until a DB match is found):
#   B  = asset_tag
#   BF = hostname (primary)
#   K  = hostname (alternate 1)
#   AS = hostname (alternate 2)
#
# Fields updated on match:
#   A  → location
#   C  → quad
#   D  → rack
#   E  → shelf

FIELDS_TO_UPDATE = ["location", "quad", "rack", "shelf"]


def col_letter_to_index(col: str) -> int:
    """Convert an Excel column letter (e.g. 'A', 'BF') to a 0-based integer index."""
    col, index = col.upper(), 0
    for ch in col:
        index = index * 26 + (ord(ch) - ord("A") + 1)
    return index - 1


# Pre-compute all column indices
COL = {letter: col_letter_to_index(letter) for letter in ("A", "B", "C", "D", "E", "K", "AS", "BF")}


class Command(BaseCommand):
    help = (
        "Read host data from an Excel sheet and update Host records "
        "(location, rack, quad, shelf). Matches DB records by asset_tag (col B), "
        "or hostname in cols BF, K, or AS — whichever finds a match first."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=None,
            help=(
                f"Path to the Excel file. "
                f"Defaults to '{EXCEL_FILENAME}' in the same directory as this script."
            ),
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simulate without saving to DB",
        )
        parser.add_argument(
            "--log-file",
            default=None,
            help=(
                "Path to write the unmatched rows log. "
                "Defaults to 'unmatched_hosts_<timestamp>.log' next to the Excel file."
            ),
        )

    def handle(self, *args, **opts):
        dry = opts["dry_run"]

        # ── Resolve Excel file path ───────────────────────────────────────────
        if opts["file"]:
            excel_path = Path(opts["file"])
        else:
            excel_path = Path(__file__).resolve().parent / EXCEL_FILENAME

        if not excel_path.exists():
            self.stdout.write(self.style.ERROR(f"Excel file not found: {excel_path}"))
            return

        # ── Resolve log file path ─────────────────────────────────────────────
        if opts["log_file"]:
            log_path = Path(opts["log_file"])
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_path = excel_path.parent / f"unmatched_hosts_{timestamp}.log"

        self.stdout.write(f"Reading Excel : {excel_path}")
        self.stdout.write(f"Unmatched log : {log_path}")
        if dry:
            self.stdout.write(self.style.WARNING("DRY RUN - no database writes"))

        # ── Load workbook ─────────────────────────────────────────────────────
        try:
            import openpyxl
        except ImportError:
            self.stdout.write(self.style.ERROR("openpyxl not installed. Run: pip install openpyxl"))
            return

        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        except Exception as exc:
            self.stdout.write(self.style.ERROR(f"Failed to open Excel file: {exc}"))
            return

        if SHEET_NAME not in wb.sheetnames:
            self.stdout.write(
                self.style.ERROR(
                    f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}"
                )
            )
            return

        ws = wb[SHEET_NAME]

        updates   = []
        unmatched = []   # rows where no lookup matched
        matched   = 0
        skipped   = 0    # rows where every lookup column is blank
        not_found = 0    # rows with values but no DB match

        def cell_val(row, idx):
            """Safely read a cell value; return stripped str or None."""
            try:
                val = row[idx]
                return str(val).strip() if val is not None else None
            except IndexError:
                return None

        # ── Iterate rows (skip header row 1) ──────────────────────────────────
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            # ── Read all lookup columns ───────────────────────────────────────
            xl_asset_tag = cell_val(row, COL["B"])    # asset_tag  – col B
            hn_bf        = cell_val(row, COL["BF"])   # hostname   – col BF
            hn_k         = cell_val(row, COL["K"])    # hostname   – col K
            hn_as        = cell_val(row, COL["AS"])   # hostname   – col AS

            # Skip row entirely if every lookup column is blank
            if not xl_asset_tag and not hn_bf and not hn_k and not hn_as:
                skipped += 1
                continue

            # ── Read fields to update ─────────────────────────────────────────
            xl_location = cell_val(row, COL["A"])
            xl_quad     = cell_val(row, COL["C"])
            xl_rack     = cell_val(row, COL["D"])
            xl_shelf    = cell_val(row, COL["E"])

            # ── Try each lookup in priority order: B → BF → K → AS ───────────
            #    First one that returns a DB hit wins.
            existing    = Host.objects.none()
            matched_col = None
            matched_val = None

            lookup_order = [
                ("asset_tag (B)", "asset_tag__iexact",    xl_asset_tag),
                ("hostname (BF)", "hostname__icontains",  hn_bf),
                ("hostname (K)",  "hostname__icontains",  hn_k),
                ("hostname (AS)", "hostname__icontains",  hn_as),
            ]

            for col_label, lookup_field, value in lookup_order:
                if value:
                    qs = Host.objects.filter(**{lookup_field: value})
                    if qs.exists():
                        existing    = qs
                        matched_col = col_label
                        matched_val = value
                        break

            # ── No match found ────────────────────────────────────────────────
            if not existing.exists():
                not_found += 1
                unmatched.append((row_num, xl_asset_tag, hn_bf, hn_k, hn_as))
                self.stdout.write(
                    self.style.WARNING(
                        f"Row {row_num:5d}: no DB match  "
                        f"[asset_tag={xl_asset_tag!r}  BF={hn_bf!r}  "
                        f"K={hn_k!r}  AS={hn_as!r}] – skipped."
                    )
                )
                continue

            # ── Apply field updates ───────────────────────────────────────────
            for host in existing:
                host.location = xl_location
                host.quad     = xl_quad
                host.rack     = xl_rack
                host.shelf    = xl_shelf
                updates.append(host)

            matched += len(existing)
            self.stdout.write(
                self.style.SUCCESS(
                    f"Row {row_num:5d}: Matched {len(existing)} host(s) "
                    f"via {matched_col} → {matched_val}"
                )
            )

        wb.close()

        # ── Write unmatched log ───────────────────────────────────────────────
        try:
            with open(log_path, "w", encoding="utf-8") as lf:
                lf.write("Unmatched Hosts Log\n")
                lf.write("Generated  : " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n")
                lf.write("Excel file : " + str(excel_path) + "\n")
                lf.write("Dry run    : " + str(dry) + "\n")
                lf.write("=" * 90 + "\n")
                if unmatched:
                    lf.write(
                        f"{'Row':>6}  {'Asset Tag (B)':<20}  {'Hostname BF':<25}  "
                        f"{'Hostname K':<25}  Hostname AS\n"
                    )
                    lf.write("-" * 90 + "\n")
                    for r, at, bf, k, a in unmatched:
                        lf.write(
                            f"{r:>6}  {(at or '-'):<20}  {(bf or '-'):<25}  "
                            f"{(k or '-'):<25}  {a or '-'}\n"
                        )
                else:
                    lf.write("All rows matched successfully – no unmatched entries.\n")
                lf.write("=" * 90 + "\n")
                lf.write(f"Total unmatched: {len(unmatched)}\n")
            self.stdout.write(self.style.SUCCESS(f"Log written  : {log_path}"))
        except OSError as exc:
            self.stdout.write(self.style.ERROR(f"Could not write log file: {exc}"))

        # ── Summary ───────────────────────────────────────────────────────────
        self.stdout.write("\n" + "═" * 70)
        self.stdout.write("Import Summary:")
        self.stdout.write(f"  Matched / to update    : {matched}")
        self.stdout.write(f"  Not found in DB        : {not_found}")
        self.stdout.write(f"  Skipped (all blank)    : {skipped}")
        self.stdout.write(f"  Total rows processed   : {matched + not_found + skipped}")
        self.stdout.write("═" * 70 + "\n")

        if dry:
            self.stdout.write(self.style.SUCCESS("Dry run complete — no changes saved."))
            return

        if updates:
            with transaction.atomic():
                Host.objects.bulk_update(updates, FIELDS_TO_UPDATE, batch_size=1000)
            self.stdout.write(
                self.style.SUCCESS(f"Updated {len(updates)} host record(s) successfully.")
            )
        else:
            self.stdout.write(self.style.WARNING("No matching hosts found – nothing to update."))